#!/usr/bin/env python3
"""Script to create test Excel files for converter testing."""

import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def create_excel_test_files():
    """Create comprehensive Excel test files."""
    # Create test directory if it doesn't exist
    test_dir = os.path.dirname(__file__)

    # 1. Sales Data Excel File
    sales_data = {
        "Product": ["Widget A", "Widget B", "Widget C", "Widget D"],
        "Q1 Sales": [1000, 800, 1500, 2000],
        "Q2 Sales": [1200, 900, 1600, 2100],
        "Q3 Sales": [1100, 850, 1550, 2050],
        "Q4 Sales": [1300, 950, 1700, 2200],
        "Total": [4600, 3500, 6350, 8350],
    }

    employee_data = {
        "Name": ["John Doe", "Jane Smith", "Bob Johnson", "Alice Brown"],
        "Department": ["Engineering", "Marketing", "Sales", "HR"],
        "Salary": [75000, 65000, 70000, 60000],
        "Hire Date": ["2020-01-15", "2019-03-22", "2021-06-10", "2018-11-05"],
    }

    financial_data = {
        "Metric": ["Revenue", "Expenses", "Profit", "Growth Rate"],
        "Value": [22800, 18240, 4560, 15],
        "Percentage": ["100%", "80%", "20%", "-"],
    }

    # Create Excel file with multiple sheets
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Sales Data Sheet
    ws_sales = wb.create_sheet("Sales Data")
    df_sales = pd.DataFrame(sales_data)
    for r in dataframe_to_rows(df_sales, index=False, header=True):
        ws_sales.append(r)

    # Style the header
    for cell in ws_sales[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

    # Employee Data Sheet
    ws_emp = wb.create_sheet("Employee Data")
    df_emp = pd.DataFrame(employee_data)
    for r in dataframe_to_rows(df_emp, index=False, header=True):
        ws_emp.append(r)

    # Style the header
    for cell in ws_emp[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

    # Financial Summary Sheet
    ws_fin = wb.create_sheet("Financial Summary")
    df_fin = pd.DataFrame(financial_data)
    for r in dataframe_to_rows(df_fin, index=False, header=True):
        ws_fin.append(r)

    # Style the header
    for cell in ws_fin[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

    # Save the file
    excel_path = os.path.join(test_dir, "test_excel_sample.xlsx")
    wb.save(excel_path)
    print(f"Created Excel file: {excel_path}")

    # 2. Simple CSV file
    csv_data = {
        "Name": ["Alice", "Bob", "Charlie", "Diana"],
        "Age": [25, 30, 35, 28],
        "City": ["New York", "London", "Tokyo", "Paris"],
        "Salary": [50000, 60000, 70000, 55000],
    }

    df_csv = pd.DataFrame(csv_data)
    csv_path = os.path.join(test_dir, "test_csv_sample.csv")
    df_csv.to_csv(csv_path, index=False)
    print(f"Created CSV file: {csv_path}")

    # 3. Complex CSV with special characters
    complex_csv_data = {
        "Product Name": [
            'Widget "Special" Edition',
            "Gadget's Delight",
            "Tool & Accessory Set",
        ],
        "Price": ["$29.99", "$49.99", "$79.99"],
        "Description": [
            'High-quality widget with "premium" features',
            "Gadget for all occasions",
            "Complete tool set with accessories",
        ],
        "Category": ["Electronics", "Home & Garden", "Tools"],
    }

    df_complex = pd.DataFrame(complex_csv_data)
    complex_csv_path = os.path.join(test_dir, "test_csv_complex.csv")
    df_complex.to_csv(complex_csv_path, index=False, quoting=1)  # Quote all fields
    print(f"Created complex CSV file: {complex_csv_path}")


if __name__ == "__main__":
    create_excel_test_files()
