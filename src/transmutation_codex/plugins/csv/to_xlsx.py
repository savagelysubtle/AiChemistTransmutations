"""CSV to Excel converter.

This module converts CSV files to Excel format with formatting and styling.
"""

from pathlib import Path
from typing import Any

try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pd = None

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None

from transmutation_codex.core import (
    ConversionEvent,
    EventTypes,
    check_feature_access,
    check_file_size_limit,
    complete_operation,
    get_log_manager,
    publish,
    raise_conversion_error,
    record_conversion_attempt,
    start_operation,
    update_progress,
)
from transmutation_codex.core.decorators import converter

# Setup logger
logger = get_log_manager().get_converter_logger("csv2xlsx")


@converter(
    source_format="csv",
    target_format="xlsx",
    description="Convert CSV to formatted Excel spreadsheet",
    required_dependencies=["pandas", "openpyxl"],
    priority=10,
    version="1.0.0",
)
def convert_csv_to_xlsx(
    input_path: str | Path,
    output_path: str | Path | None = None,
    **kwargs: Any,
) -> Path:
    """Convert CSV file to Excel format.

    This function converts CSV files to Excel format (.xlsx) with formatting,
    styling, and proper data types.

    Args:
        input_path: Path to input CSV file
        output_path: Path for output Excel file (auto-generated if None)
        **kwargs: Additional options:
            - encoding: CSV encoding (default: 'utf-8')
            - separator: CSV separator (default: ',')
            - sheet_name: Excel sheet name (default: 'Sheet1')
            - include_header: Include header row (default: True)
            - format_headers: Format header row (default: True)
            - auto_width: Auto-adjust column widths (default: True)

    Returns:
        Path: Path to generated Excel file

    Raises:
        ConversionError: If conversion fails
        ValidationError: If input validation fails
    """
    # Validate dependencies
    if not PANDAS_AVAILABLE:
        raise_conversion_error("pandas is required for CSV conversion")
    if not OPENPYXL_AVAILABLE:
        raise_conversion_error("openpyxl is required for Excel generation")

    # Start operation
    operation = start_operation(
        "conversion", f"Converting CSV to Excel: {Path(input_path).name}"
    )

    try:
        # Check licensing and file size
        check_feature_access("csv2xlsx")
        check_file_size_limit(input_path, max_size_mb=100)
        record_conversion_attempt("csv2xlsx")

        # Convert paths
        input_path = Path(input_path)
        if output_path is None:
            output_path = input_path.with_suffix(".xlsx")
        else:
            output_path = Path(output_path)

        logger.info(f"Converting CSV to Excel: {input_path} -> {output_path}")

        # Parse options
        encoding = kwargs.get("encoding", "utf-8")
        separator = kwargs.get("separator", ",")
        sheet_name = kwargs.get("sheet_name", "Sheet1")
        include_header = kwargs.get("include_header", True)
        format_headers = kwargs.get("format_headers", True)
        auto_width = kwargs.get("auto_width", True)

        update_progress(operation.id, 10, "Reading CSV file...")

        # Read CSV file
        try:
            df = pd.read_csv(
                input_path,
                encoding=encoding,
                sep=separator,
                header=0 if include_header else None,
            )
        except Exception as e:
            raise_conversion_error(f"Failed to read CSV file: {e}")

        update_progress(operation.id, 30, "Creating Excel file...")

        # Create Excel file
        try:
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            raise_conversion_error(f"Failed to create Excel file: {e}")

        update_progress(operation.id, 60, "Applying formatting...")

        # Apply formatting if requested
        if format_headers or auto_width:
            try:
                workbook = openpyxl.load_workbook(output_path)
                worksheet = workbook[sheet_name]

                if format_headers and include_header:
                    # Format header row
                    header_fill = PatternFill(
                        start_color="366092", end_color="366092", fill_type="solid"
                    )
                    header_font = Font(color="FFFFFF", bold=True)

                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font

                if auto_width:
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter

                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass

                        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                        worksheet.column_dimensions[
                            column_letter
                        ].width = adjusted_width

                workbook.save(output_path)
                workbook.close()

            except Exception as e:
                logger.warning(f"Failed to apply formatting: {e}")
                # Continue without formatting

        update_progress(operation.id, 90, "Conversion completed...")

        # Publish success event
        publish(
            ConversionEvent(
                event_type=EventTypes.CONVERSION_COMPLETED,
                input_file=str(input_path),
                output_file=str(output_path),
                conversion_type="conversion",
            )
        )

        complete_operation(operation.id, {"output_path": str(output_path)})
        logger.info(f"CSV to Excel conversion completed: {output_path}")

        return output_path

    except Exception as e:
        logger.exception(f"CSV to Excel conversion failed: {e}")
        publish(
            ConversionEvent(
                event_type=EventTypes.CONVERSION_FAILED,
                input_file=str(input_path),
                conversion_type="conversion",
            )
        )
        raise_conversion_error(f"Conversion failed: {e}")

