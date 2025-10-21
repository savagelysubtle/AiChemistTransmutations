"""Excel to PDF converter.

This module converts Excel spreadsheets to PDF format, preserving formatting,
charts, and layout.
"""

from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None

try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pd = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )
    from reportlab.platypus.flowables import PageBreak

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from transmutation_codex.core import (
    ConversionEvent,
    EventTypes,
    check_feature_access,
    check_file_size_limit,
    complete_operation,
    get_log_manager,
    publish,
    raise_conversion_error,
    raise_validation_error,
    record_conversion_attempt,
    start_operation,
    update_progress,
)
from transmutation_codex.core.decorators import converter

# Setup logger
logger = get_log_manager().get_converter_logger("xlsx2pdf")


@converter(
    source_format="xlsx",
    target_format="pdf",
    description="Convert Excel to PDF with formatting preservation",
    required_dependencies=["openpyxl", "pandas", "reportlab"],
    priority=10,
    version="1.0.0",
)
def convert_xlsx_to_pdf(
    input_path: str | Path,
    output_path: str | Path | None = None,
    **kwargs: Any,
) -> Path:
    """Convert Excel spreadsheet to PDF format.

    This function converts Excel files (.xlsx, .xls) to PDF format while
    preserving formatting, charts, and layout. It handles multiple sheets
    and creates a professional-looking PDF output.

    Args:
        input_path: Path to input Excel file
        output_path: Path for output PDF file (auto-generated if None)
        **kwargs: Additional options:
            - include_charts: Include charts in PDF (default: True)
            - page_size: Page size ('A4', 'Letter', default: 'A4')
            - orientation: Page orientation ('portrait', 'landscape', default: 'portrait')
            - include_gridlines: Show gridlines (default: True)
            - include_headers: Include headers/footers (default: True)
            - max_rows_per_page: Maximum rows per page (default: 50)
            - font_size: Font size for PDF (default: 8)

    Returns:
        Path: Path to generated PDF file

    Raises:
        ConversionError: If conversion fails
        ValidationError: If input validation fails
    """
    # Validate dependencies
    if not OPENPYXL_AVAILABLE:
        raise_conversion_error("openpyxl is required for Excel conversion")
    if not PANDAS_AVAILABLE:
        raise_conversion_error("pandas is required for Excel conversion")
    if not REPORTLAB_AVAILABLE:
        raise_conversion_error("reportlab is required for PDF generation")

    # Start operation
    operation_id = start_operation(
        f"Converting Excel to PDF: {Path(input_path).name}", total_steps=100
    )

    try:
        # Check licensing and file size
        check_feature_access("xlsx2pdf")
        check_file_size_limit(input_path)
        record_conversion_attempt("xlsx2pdf", str(input_path))

        # Convert paths
        input_path = Path(input_path)
        if output_path is None:
            output_path = input_path.with_suffix(".pdf")
        else:
            output_path = Path(output_path)

        logger.info(f"Converting Excel to PDF: {input_path} -> {output_path}")

        # Parse options
        include_charts = kwargs.get("include_charts", True)
        page_size = kwargs.get("page_size", "A4")
        orientation = kwargs.get("orientation", "portrait")
        include_gridlines = kwargs.get("include_gridlines", True)
        include_headers = kwargs.get("include_headers", True)
        max_rows_per_page = kwargs.get("max_rows_per_page", 50)
        font_size = kwargs.get("font_size", 8)

        # Validate page size
        if page_size.lower() == "a4":
            pagesize = A4
        elif page_size.lower() == "letter":
            pagesize = letter
        else:
            raise_validation_error(f"Invalid page size: {page_size}")

        # Validate orientation
        if orientation.lower() == "landscape":
            pagesize = (pagesize[1], pagesize[0])  # Swap width/height

        update_progress(operation_id, 10, "Loading Excel file...")

        # Load Excel file
        try:
            workbook = openpyxl.load_workbook(input_path, data_only=True)
        except Exception as e:
            raise_conversion_error(f"Failed to load Excel file: {e}")

        update_progress(operation_id, 20, "Processing worksheets...")

        # Create PDF document
        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=pagesize,
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )

        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=14,
            spaceAfter=12,
            alignment=1,  # Center alignment
        )

        story = []

        # Process each worksheet
        sheet_names = workbook.sheetnames
        total_sheets = len(sheet_names)

        for sheet_idx, sheet_name in enumerate(sheet_names):
            logger.info(f"Processing sheet: {sheet_name}")
            update_progress(
                operation_id,
                20 + (sheet_idx / total_sheets) * 60,
                f"Processing sheet: {sheet_name}",
            )

            worksheet = workbook[sheet_name]

            # Add sheet title
            if include_headers:
                story.append(Paragraph(f"<b>{sheet_name}</b>", title_style))
                story.append(Spacer(1, 12))

            # Get data from worksheet
            data = []
            max_col = worksheet.max_column
            max_row = worksheet.max_row

            # Process data in chunks to avoid memory issues
            chunk_size = max_rows_per_page
            for start_row in range(1, max_row + 1, chunk_size):
                end_row = min(start_row + chunk_size - 1, max_row)

                # Get chunk data
                chunk_data = []
                for row in range(start_row, end_row + 1):
                    row_data = []
                    for col in range(1, max_col + 1):
                        cell = worksheet.cell(row=row, column=col)
                        value = cell.value

                        # Handle different data types
                        if value is None:
                            row_data.append("")
                        elif isinstance(value, (int, float)):
                            row_data.append(str(value))
                        else:
                            row_data.append(str(value))

                    chunk_data.append(row_data)

                if chunk_data:
                    # Create table
                    table = Table(chunk_data)

                    # Style the table
                    table_style = [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), font_size),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.black)
                        if include_gridlines
                        else None,
                    ]

                    # Remove None values
                    table_style = [style for style in table_style if style is not None]

                    table.setStyle(TableStyle(table_style))
                    story.append(table)

                    # Add page break if not the last chunk
                    if end_row < max_row:
                        story.append(PageBreak())

            # Add page break between sheets (except for the last sheet)
            if sheet_idx < total_sheets - 1:
                story.append(PageBreak())

        update_progress(operation_id, 90, "Generating PDF...")

        # Build PDF
        doc.build(story)

        # Publish success event
        publish(
            ConversionEvent(
                event_type=EventTypes.CONVERSION_COMPLETED,
                input_file=str(input_path),
                output_file=str(output_path),
                conversion_type="xlsx2pdf",
            )
        )

        complete_operation(operation_id, {"output_path": str(output_path)})
        logger.info(f"Excel to PDF conversion completed: {output_path}")

        return output_path

    except Exception as e:
        logger.exception(f"Excel to PDF conversion failed: {e}")
        publish(
            ConversionEvent(
                event_type=EventTypes.CONVERSION_FAILED,
                input_file=str(input_path),
                conversion_type="xlsx2pdf",
            )
        )
        raise_conversion_error(f"Excel to PDF conversion failed: {e}")
